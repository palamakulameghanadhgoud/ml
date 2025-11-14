import openpyxl
import os


file_path = r"D:\ml projects\ML PROJECT VOLUNTEERS .xlsx"
workbook = openpyxl.load_workbook(file_path)
sheet = workbook.active


base_folder_path = r"D:\ml projects\volunteers"


folder_col = sheet.max_column + 1
imgdata_col = folder_col + 1
sheet.cell(row=1, column=folder_col).value = "Folder Path"
sheet.cell(row=1, column=imgdata_col).value = "Img Data"

for row in range(2, sheet.max_row + 1):
    university_id = str(sheet.cell(row=row, column=2).value)  
    
   
    folder_path = os.path.join(base_folder_path, university_id)
    sheet.cell(row=row, column=folder_col).value = folder_path


    sheet.cell(row=row, column=imgdata_col).value = f"{university_id}_photo.jpg"

workbook.save(file_path)
print("Excel updated with Folder Path and Img Data columns!")